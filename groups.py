from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from openpyxl import load_workbook, Workbook, formatting, styles
from openpyxl.worksheet.worksheet import Worksheet


from colorama import init as colorama_init
from colorama import Fore, Back
from colorama import Style


colorama_init()

SUBJECTS = [
    # name      num of groups
    ('SK',      4),
    ('CPS',     4),
    ('SS',      4),
    ('BST',     4),
    ('JPWP',    4),
    ('WWW+L',   4),
    ('WWW+S',   4),
]

SUBJECTS_NAME = [name for name, _ in SUBJECTS]
"""Order as in excel."""

@dataclass
class Person:
    sem: Semester = field(repr=False)
    name: str
    index: int
    
    gr_status: str = field(repr=False)
    change: str = field(repr=False)
    why: str = field(repr=False)
    who: str = field(repr=False)
    more: str = field(repr=False)

    friend_group: FriendGroup = field(repr=False, default=None)
    pref_gr: dict[LabGroupStatus, int] = field(default_factory=dict, repr=False)

    def __post_init__(self) -> None:
        for num, status in enumerate(self.gr_status.strip().split(";"), start=1):
            self.pref_gr[LabGroupStatus(status)] = num 

    def chgr(self, sub: Subject, dstgr: int):
        """Change group of a subject."""
        self.sem.db[self][sub] = dstgr

    def print_groups(self):
        for subj, gr in self.sem.db[self].items():
            print(f"{subj}\t{gr}")

    def __hash__(self) -> int:
        return id(self)


@dataclass
class FriendGroup:
    sem: Semester
    members: list[Person]

    def match(self):
        all_subjects: dict[Subject, dict[Person, SubjectGroup]] = {s:{} for s in self.sem.subjects}
        for person in self.members:
            for subj, group in self.sem.db[person].items():
                all_subjects[subj][person] = group

        for subj, person_group in all_subjects.items():
            s = set(list(person_group.values()))
            print(f"Match for {subj.name} is `100/len(s)`={100/len(s)}%")


class LabGroupStatus(Enum):
    best = "najlepsza"
    fine = "ujdzie"
    plsno = "prosze nie"
    nooo = "absolutnie nie"


@dataclass
class SubjectGroup:
    """Group per subject."""
    subject: Subject = field(repr=False)
    group_num: int
    name: str = ""
    datetime: str = field(repr=False, default="")

    members: list[Person] = field(init=False, default_factory=list, repr=False)

    def __hash__(self) -> int:
        return id(self)


@dataclass
class Subject:
    sem: Semester = field(repr=False)
    name: str
    lab_groups: dict[int, SubjectGroup] = field(init=False, default_factory=dict, repr=False)
    num_groups: int = field(default=4, repr=False)

    def __post_init__(self):
        for i in range(1, self.num_groups + 1):
            self.lab_groups[i] = SubjectGroup(subject=self, group_num=i, name=self.name)

    def mvper(self, p: Person, dstgr: int) -> SubjectGroup:
        if not dstgr in self.lab_groups:
            raise ValueError(f"Invalid group number = {dstgr}")

        self.sem.db[p][self] = self[dstgr]

    def __getitem__(self, key: int) -> SubjectGroup:
        return self.lab_groups[key]

    def __hash__(self) -> int:
        return id(self)

@dataclass
class Semester:
    excel_path: str
    wb: Workbook = field(init=False, repr=False)
    subjects: dict[str, Subject] = field(init=False, default_factory=dict)
    db: dict[Person, dict[Subject, SubjectGroup]] = field(init=False, default_factory=dict)
    
    def __post_init__(self) -> None:
        for name, num_groups in SUBJECTS:
            self.subjects[name] = Subject(self, name, num_groups=num_groups)
        self.wb = load_workbook(self.excel_path)
        self._people_from_excel()
        self._groups_from_excel()

    def new_person(self, *args, **kwargs) -> Person:
        p = Person(self, *args, **kwargs)
        self.db[p] = { s:None for s in self.subjects.values() }
        return p

    def get_person(self, index: int) -> Person:
        for p in self.db.keys():
            if p.index == index:
                return p
        return None

    def _people_from_excel(self) -> None:
        forms: Worksheet = self.wb['Liczba odpowiedzi 1']
        
        for row in forms.iter_rows(min_row=57, max_row=111,
                                   min_col=1, max_col=11,
                                   values_only=True):
            _, name, index, gr1, gr2, gr3, gr4, change, why, who, more = self.row_float2int(row)
            self.new_person(name, index, ";".join([gr1, gr2, gr3, gr4]), change, why, who, more)

    def _groups_from_excel(self) -> None:
        groups: Worksheet = self.wb['podejscie 2']
        for row in groups.iter_rows(min_row=2, max_row=66,
                                    min_col=1, max_col=9,
                                    values_only=True):
            if row[0] == None:
                continue # this row is empty, and its purpuse is better readbility

            index, name, sk, cps, ss, bst, jpwp, wwwL, wwwS = self.row_float2int(row)
            p = self.get_person(index)
            if p is None:
                # raise RuntimeError(f'{index=} was not found!')
                print(f'{index=} was not found!')
                continue

            self.subjects['SK'   ].mvper(p, sk)
            self.subjects['CPS'  ].mvper(p, cps)
            self.subjects['SS'   ].mvper(p, ss)
            self.subjects['BST'  ].mvper(p, bst)
            self.subjects['JPWP' ].mvper(p, jpwp)
            self.subjects['WWW+L'].mvper(p, wwwL)
            self.subjects['WWW+S'].mvper(p, wwwS)

    def _dump_groups_to_excel(self):
        shname = 'groups' # sheet name
        if not shname in self.wb:
            self.wb.create_sheet(shname)
        s = self.wb[shname]
        s.append(['index', 'imie i nazwisko'] + SUBJECTS_NAME[:])
        for p in self.db.keys():
            s.append([p.index, p.name]
                      + [
                            # person -> subject -> this persons group -> group number
                            self.db[p][self.subjects[name]].group_num
                            for name in SUBJECTS_NAME
                    ])
        
        cell_range = "A1:I99999"
        for gr_num, color in {
                                1: 'ab82ba',    # grayish purple
                                2: 'b8edb7',    # light green
                                3: 'fadd5f',    # yellow
                                4: 'e06158',    # red
                            }.items():
            s.conditional_formatting.add(cell_range, formatting.rule.CellIsRule(operator='==',
                                                                                formula=[str(gr_num)],
                                                                                fill=styles.PatternFill(start_color=color,
                                                                                                        end_color=color,
                                                                                                        fill_type='solid')))
            

    def check_group_confilts(self):
        CPS  = self.subjects['CPS'  ]
        SS   = self.subjects['SS'   ]
        BST  = self.subjects['BST'  ]
        JPWP = self.subjects['JPWP' ]
        WWWL = self.subjects['WWW+L']
        
        conflicts = [
            [SS[2],     CPS[3]],
            [WWWL[3],   CPS[4]],
            [CPS[2],    WWWL[4]],
            [CPS[1],    JPWP[4]],
            [JPWP[1],   SS[3],      BST[4]],
            [BST[1],    JPWP[2],    SS[3]],
            [WWWL[1],   BST[2],     JPWP[3],    SS[4]],
            [WWWL[2],   BST[3],     SS[4]],
        ]
        count_conflicts = 0
        for p, subj_grp in self.db.items():
            pgrps = set( list( subj_grp.values() ))

            for c in conflicts:
                inter = pgrps.intersection(set(c))

                if len(inter) > 1:
                    count_conflicts += 1
                    print(f'{Fore.RED}CONFILCT DETECTED!{Style.RESET_ALL}',
                                p,
                                f'{Back.RED}{inter}{Style.RESET_ALL}',
                          sep="\n\t")

        if count_conflicts == 0:
            print(f'{Fore.GREEN}No conflicts detected!{Style.RESET_ALL}')
        else:
            print(f'{Fore.RED}Conflicts detected: {count_conflicts}.{Style.RESET_ALL}')

    def save(self):
        self._dump_groups_to_excel()
        self.wb.save(self.excel_path)
        self.wb.close()

    @staticmethod
    def row_float2int(row) -> list:
        return [(int(v) if isinstance(v, float) else v) for v in row]


if __name__ == "__main__":
    s =  Semester(r"C:\Users\quatr\Downloads\Zapisy do grup (Odpowiedzi) (4).xlsx")
    s.check_group_confilts()
    s.save()


